from openpyxl import Workbook, load_workbook
from datetime import datetime

wb = load_workbook(filename="Employees.xlsx")
sheetnames = wb.sheetnames
headers = ["emp_id", "first_name", "last_name", "hire_date"]

combined = []
# Loop through all sheetnames and and combine all into one array
for sheet in sheetnames:
	page = wb[sheet]
	first_row = True
	for item in page.rows:
		# use the value instead of object
		values = []
		for val in item:
			values.append(val.value)

		if type(values[0]) == int:
			combined.append(values)

combined.sort(reverse = True, key=lambda x: x[3])

empty = []
for row in combined:
	empty_row = []
	for item in row:
		if type(item) == datetime:
			item_str = item.strftime("%m/%d/%Y")
			empty_row.append(item_str)

		else:
			empty_row.append(item)
	empty.append(empty_row)


new_wb = Workbook()
page = new_wb.active

page.append(headers)
for item in empty:
	page.append(item)

new_wb.save(filename="Combined.xlsx")