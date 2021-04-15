import statistics
from openpyxl import Workbook, load_workbook

headers = ["Sum", "Mean", "Max", "Min"]

wb = load_workbook(filename="Coalmines2019.xlsx")
ws = wb.active

dictio = []
production_row = 3
for item in ws.rows:
	dictio.append(item[production_row].value)

dictio.pop(0)

data = [
	sum(dictio),
	statistics.mean(dictio),
	min(dictio),
	max(dictio)
]


new_wb = Workbook()
page = new_wb.active

page.append(headers)
page.append(data)

new_wb.save(filename="Results.xlsx")